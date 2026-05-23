#!/usr/bin/env python3
"""
Excel 수식 방지 기능 테스트
"""
from pathlib import Path
from openpyxl import Workbook, load_workbook

# 테스트 데이터 (수식이 포함된 텍스트)
test_data = [
    {"번호": "1", "질문": "다음 중 올바른 것은?", "문항1": '=CHOOSE(MID(A1,8,1), "남","여")', "문항2": "일반 텍스트", "정답": "1"},
    {"번호": "2", "질문": "계산 결과는?", "문항1": "=SUM(A1:A10)", "문항2": "+1+2+3", "정답": "2"},
    {"번호": "3", "질문": "TRUE/FALSE?", "문항1": "=IF(A1>0, TRUE, FALSE)", "문항2": "@VLOOKUP(A1, B:C, 2)", "정답": "1"},
]

headers = ["번호", "질문", "문항1", "문항2", "정답"]

print("="*80)
print("Excel 수식 방지 기능 테스트")
print("="*80)

# 1. Excel 파일 작성 (수식 방지 적용)
test_file = Path("test_formula.xlsx")
workbook = Workbook()
sheet = workbook.active
sheet.title = "테스트"
sheet.append(headers)

print("\n✍️  Excel 파일 작성 중...")
for row in test_data:
    row_values = []
    for key in headers:
        value = row.get(key, "")
        # 수식 방지: =, +, -, @로 시작하는 텍스트는 앞에 '를 붙여서 텍스트로 저장
        if value and isinstance(value, str) and value[0] in ('=', '+', '-', '@'):
            print(f"   수식 감지: '{value}' -> '''{value}'")
            value = "'" + value
        row_values.append(value)
    sheet.append(row_values)

workbook.save(test_file)
print(f"✅ 저장 완료: {test_file}")

# 2. Excel 파일 읽기 (작은따옴표 제거)
print("\n📖 Excel 파일 읽기 중...")
workbook = load_workbook(test_file)
sheet = workbook.active
rows = list(sheet.iter_rows(values_only=True))

read_headers = rows[0]
print(f"   헤더: {read_headers}")

for idx, values in enumerate(rows[1:], start=1):
    print(f"\n   문제 {idx}:")
    for header, value in zip(read_headers, values):
        if value:
            value_str = str(value)
            # 작은따옴표 제거
            if value_str.startswith("'") and len(value_str) > 1:
                original = value_str
                value_str = value_str[1:]
                print(f"      {header}: '{original}' -> '{value_str}'")
            else:
                print(f"      {header}: '{value_str}'")

# 3. 검증
print("\n" + "="*80)
print("검증 결과")
print("="*80)

success = True
for idx, values in enumerate(rows[1:], start=1):
    question_values = dict(zip(read_headers, values))
    expected = test_data[idx-1]
    
    for key in ["문항1", "문항2"]:
        actual = str(question_values.get(key, ""))
        # 작은따옴표 제거
        if actual.startswith("'"):
            actual = actual[1:]
        
        expected_value = expected.get(key, "")
        
        if actual == expected_value:
            print(f"✅ 문제 {idx} - {key}: '{actual}'")
        else:
            print(f"❌ 문제 {idx} - {key}: 기대값='{expected_value}', 실제값='{actual}'")
            success = False

# 정리
test_file.unlink()
print(f"\n🗑️  테스트 파일 삭제: {test_file}")

if success:
    print("\n🎉 모든 테스트 통과!")
else:
    print("\n⚠️  일부 테스트 실패")
