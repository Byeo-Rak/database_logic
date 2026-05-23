#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import time
from pathlib import Path
from typing import Any
from urllib import error, request

from openpyxl import load_workbook
from target_exam_files import TARGET_PDF_FILES, pdf_to_excel_path  # pyright: ignore[reportMissingImports]


OPENAI_CHAT_COMPLETIONS_URL = "https://api.openai.com/v1/chat/completions"
DEFAULT_MODEL = "gpt-5.3-chat-latest"
ERROR_MESSAGE = "ai 해설 오류"
FATAL_API_CODES = {"model_not_found", "invalid_api_key", "insufficient_quota"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Excel 문항(질문+문항1~4)만으로 GPT가 답/해설을 생성하고, "
            "정답 일치 시 해설을 기록합니다."
        )
    )
    parser.add_argument(
        "--excel",
        default="",
        help="입력 Excel 파일 경로 (생략 시 코드의 TARGET_PDF_FILES 전체 대상)",
    )
    parser.add_argument(
        "--env-file",
        default=".env",
        help="API 키가 저장된 .env 파일 경로 (기본: ./.env)",
    )
    parser.add_argument(
        "--model",
        default=DEFAULT_MODEL,
        help=f"OpenAI 모델명 (기본: {DEFAULT_MODEL})",
    )
    parser.add_argument(
        "--max-attempts",
        type=int,
        default=3,
        help="문항당 최대 시도 횟수 (기본: 3)",
    )
    parser.add_argument(
        "--sleep-seconds",
        type=float,
        default=0.2,
        help="API 호출 간 대기 시간(초) (기본: 0.2)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="기존 해설이 있어도 덮어씁니다.",
    )
    parser.add_argument(
        "--attempt-log",
        default="",
        help=(
            "시도 로그 JSONL 파일 경로 (기본: "
            "output/explanation_logs/{excel_stem}.jsonl)"
        ),
    )
    parser.add_argument(
        "--error-csv",
        default="",
        help=(
            "오류 문항 CSV 경로 (기본: "
            "output/explanation_logs/{excel_stem}_failed_rows.csv)"
        ),
    )
    parser.add_argument(
        "--from-target-files",
        action="store_true",
        help="(호환 옵션) 코드에 정의된 TARGET_PDF_FILES 목록의 Excel을 일괄 처리합니다.",
    )
    return parser.parse_args()


def load_api_key(env_file: Path) -> str:
    for env_name in ("gpt_api_key", "GPT_API_KEY", "OPENAI_API_KEY"):
        value = os.getenv(env_name, "").strip()
        if value:
            return value

    if not env_file.exists():
        raise FileNotFoundError(f".env 파일을 찾지 못했습니다: {env_file}")

    key_map: dict[str, str] = {}
    for raw_line in env_file.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key_map[key.strip()] = value.strip().strip("\"'")

    for key_name in ("gpt_api_key", "GPT_API_KEY", "OPENAI_API_KEY"):
        key_value = key_map.get(key_name, "").strip()
        if key_value:
            return key_value

    raise ValueError(
        ".env에서 API 키를 찾지 못했습니다. "
        "gpt_api_key 또는 GPT_API_KEY 또는 OPENAI_API_KEY를 설정하세요."
    )


def normalize_answer(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""

    mapping = {
        "①": "1",
        "②": "2",
        "③": "3",
        "④": "4",
        "1번": "1",
        "2번": "2",
        "3번": "3",
        "4번": "4",
    }
    if text in mapping:
        return mapping[text]

    match = re.search(r"[1-4]", text)
    if match:
        return match.group(0)
    return text


def call_openai_for_solution(
    api_key: str,
    model: str,
    question: str,
    options: list[str],
) -> tuple[str, str]:
    system_prompt = (
        "너는 객관식 문제 풀이 보조 AI다. "
        "항상 한국어로 답하고 JSON만 출력한다. "
        "정답 번호는 1~4 중 하나만 사용한다."
    )
    user_prompt = (
        f"질문:\n{question}\n\n"
        f"선택지:\n"
        f"1) {options[0]}\n"
        f"2) {options[1]}\n"
        f"3) {options[2]}\n"
        f"4) {options[3]}\n\n"
        "아래 JSON 형식으로만 응답해라.\n"
        '{"predicted_answer":"1~4 중 하나","explanation":"2~5문장 해설"}'
    )

    payload = {
        "model": model,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    }

    req = request.Request(
        OPENAI_CHAT_COMPLETIONS_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )

    try:
        with request.urlopen(req, timeout=60) as resp:
            body = json.loads(resp.read().decode("utf-8"))
    except error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        api_code = ""
        try:
            parsed_detail = json.loads(detail)
            api_code = str(parsed_detail.get("error", {}).get("code", "")).strip()
        except Exception:  # noqa: BLE001
            api_code = ""
        if api_code in FATAL_API_CODES:
            raise RuntimeError(f"FATAL_API[{api_code}] {detail}") from exc
        raise RuntimeError(f"OpenAI API 오류({exc.code}): {detail}") from exc
    except error.URLError as exc:
        raise RuntimeError(f"OpenAI API 네트워크 오류: {exc}") from exc

    content = body["choices"][0]["message"]["content"]
    parsed = json.loads(content)
    predicted = normalize_answer(str(parsed.get("predicted_answer", "")))
    explanation = str(parsed.get("explanation", "")).strip()

    if predicted not in {"1", "2", "3", "4"}:
        raise ValueError(f"유효하지 않은 predicted_answer: {predicted!r}")
    if not explanation:
        raise ValueError("빈 explanation 응답")
    return predicted, explanation


def get_column_index(headers: list[str], target: str) -> int | None:
    for idx, header in enumerate(headers, start=1):
        if header.strip() == target:
            return idx
    return None


def process_excel(args: argparse.Namespace, api_key: str, excel_path: Path) -> int:
    if not excel_path.exists():
        print(f"⚠️  Excel 파일 없음: {excel_path}")
        return 0

    workbook = load_workbook(excel_path)
    sheet = workbook.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
    explanation_col = get_column_index(headers, "해설")
    if explanation_col is None:
        explanation_col = len(headers) + 1
        sheet.cell(row=1, column=explanation_col, value="해설")

    question_col = get_column_index(headers, "질문") or 2
    answer_col = get_column_index(headers, "정답") or 7
    option_cols = [
        get_column_index(headers, "문항1") or 3,
        get_column_index(headers, "문항2") or 4,
        get_column_index(headers, "문항3") or 5,
        get_column_index(headers, "문항4") or 6,
    ]
    number_col = get_column_index(headers, "번호") or 1

    default_log_dir = excel_path.parent.parent / "explanation_logs"
    default_log_path = default_log_dir / f"{excel_path.stem}.jsonl"
    attempt_log_path = (
        Path(args.attempt_log).expanduser().resolve()
        if args.attempt_log
        else default_log_path.resolve()
    )
    attempt_log_path.parent.mkdir(parents=True, exist_ok=True)

    default_error_csv = attempt_log_path.parent / f"{excel_path.stem}_failed_rows.csv"
    error_csv_path = (
        Path(args.error_csv).expanduser().resolve()
        if args.error_csv
        else default_error_csv.resolve()
    )
    error_csv_path.parent.mkdir(parents=True, exist_ok=True)

    total_rows = max(sheet.max_row - 1, 0)
    processed = 0
    success_count = 0
    fail_count = 0
    skipped_count = 0
    attempt_records: list[dict[str, Any]] = []
    failed_rows_for_csv: list[dict[str, Any]] = []
    fatal_api_errors: list[str] = []

    for row_idx in range(2, sheet.max_row + 1):
        number = str(sheet.cell(row=row_idx, column=number_col).value or "").strip()
        question = str(sheet.cell(row=row_idx, column=question_col).value or "").strip()
        correct_answer = normalize_answer(
            str(sheet.cell(row=row_idx, column=answer_col).value or "").strip()
        )
        options = [
            str(sheet.cell(row=row_idx, column=col_idx).value or "").strip()
            for col_idx in option_cols
        ]
        current_explanation = str(
            sheet.cell(row=row_idx, column=explanation_col).value or ""
        ).strip()

        if not question or not correct_answer or any(not opt for opt in options):
            skipped_count += 1
            continue
        if current_explanation and not args.force:
            skipped_count += 1
            continue

        processed += 1
        attempts: list[dict[str, Any]] = []
        final_explanation = ERROR_MESSAGE
        is_success = False

        for attempt_no in range(1, args.max_attempts + 1):
            predicted_answer = ""
            explanation = ""
            api_error = ""
            try:
                predicted_answer, explanation = call_openai_for_solution(
                    api_key=api_key,
                    model=args.model,
                    question=question,
                    options=options,
                )
            except Exception as exc:  # noqa: BLE001
                api_error = str(exc)
                if "FATAL_API[" in api_error:
                    fatal_api_errors.append(api_error)
                    attempts.append(
                        {
                            "attempt": attempt_no,
                            "predicted_answer": "",
                            "correct_answer": correct_answer,
                            "is_correct": False,
                            "explanation": "",
                            "error": api_error,
                        }
                    )
                    break

            is_correct = predicted_answer == correct_answer if predicted_answer else False
            attempts.append(
                {
                    "attempt": attempt_no,
                    "predicted_answer": predicted_answer,
                    "correct_answer": correct_answer,
                    "is_correct": is_correct,
                    "explanation": explanation,
                    "error": api_error,
                }
            )

            if is_correct and explanation:
                is_success = True
                final_explanation = explanation
                break

            if api_error and "FATAL_API[" in api_error:
                break
            time.sleep(max(0.0, args.sleep_seconds))

        if is_success:
            success_count += 1
        else:
            fail_count += 1
            failed_rows_for_csv.append(
                {
                    "번호": number,
                    "행번호": row_idx,
                    "정답": correct_answer,
                    "질문": question,
                    "실패사유": attempts[-1]["error"] if attempts else "",
                }
            )

        sheet.cell(row=row_idx, column=explanation_col, value=final_explanation)
        attempt_records.append(
            {
                "excel_path": str(excel_path),
                "row": row_idx,
                "번호": number,
                "정답": correct_answer,
                "성공여부": is_success,
                "attempts": attempts,
            }
        )
        print(
            f"[{processed}/{total_rows}] 번호={number or row_idx} "
            f"성공={is_success} 시도={len(attempts)}"
        )
        if fatal_api_errors:
            print("치명적 API 오류 감지로 파일 처리를 중단합니다.")
            break

    workbook.save(excel_path)

    with attempt_log_path.open("w", encoding="utf-8") as log_fp:
        for record in attempt_records:
            log_fp.write(json.dumps(record, ensure_ascii=False) + "\n")

    with error_csv_path.open("w", encoding="utf-8", newline="") as csv_fp:
        writer = csv.DictWriter(
            csv_fp,
            fieldnames=["번호", "행번호", "정답", "질문", "실패사유"],
        )
        writer.writeheader()
        writer.writerows(failed_rows_for_csv)

    print("-" * 80)
    print(f"Excel 저장 완료: {excel_path}")
    print(f"시도 로그 저장: {attempt_log_path}")
    print(f"실패 요약 CSV 저장: {error_csv_path}")
    print(
        "처리 요약: "
        f"처리={processed}, 성공={success_count}, 실패={fail_count}, 건너뜀={skipped_count}"
    )
    if fatal_api_errors:
        print(f"치명적 API 오류: {fatal_api_errors[0]}")
        return 1
    return 0


def resolve_excel_targets(args: argparse.Namespace, project_root: Path) -> list[Path]:
    if args.excel:
        return [Path(args.excel).expanduser().resolve()]

    output_root = project_root / "output"
    return [pdf_to_excel_path(pdf, output_root).resolve() for pdf in TARGET_PDF_FILES]


def main() -> int:
    args = parse_args()
    if args.max_attempts < 1:
        raise ValueError("--max-attempts는 1 이상이어야 합니다.")

    project_root = Path(".").resolve()
    env_file = Path(args.env_file).expanduser().resolve()
    api_key = load_api_key(env_file)
    excel_targets = resolve_excel_targets(args, project_root)

    print("=" * 80)
    print("GPT 해설 생성 시작")
    print(f"대상 파일 수: {len(excel_targets)}")
    print("=" * 80)

    success_files = 0
    fail_files = 0
    for excel_path in excel_targets:
        try:
            exit_code = process_excel(args=args, api_key=api_key, excel_path=excel_path)
            if exit_code == 0:
                success_files += 1
            else:
                fail_files += 1
        except Exception as exc:  # noqa: BLE001
            fail_files += 1
            print(f"❌ 처리 실패: {excel_path}")
            print(f"   사유: {exc}")

    print("=" * 80)
    print("GPT 해설 생성 종료")
    print(f"성공 파일: {success_files}")
    print(f"실패 파일: {fail_files}")
    print("=" * 80)
    return 0 if fail_files == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
