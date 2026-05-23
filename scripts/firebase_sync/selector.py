from __future__ import annotations

import csv
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

from .models import ExamFileSet


# CSV는 더 이상 사용하지 않음 (Excel만 사용)
PAGE_MARKER_PATTERN = re.compile(r"\[PAGE\s+(?P<page>\d+)\]")
QUESTION_NUMBER_PATTERN = re.compile(r"(?<!\d)(\d{1,3})\.\s")
IMAGE_NAME_PATTERN = re.compile(r"^page_(?P<page>\d{3})_img_(?P<idx>\d{2})")
# 새로운 Firebase Storage 스타일 이미지 경로 패턴
# 경로: images/{cert}/{company}/{round-subject}/{qno}/question-1.jpg
FIREBASE_IMAGE_PATH_PATTERN = re.compile(
    r"[/\\](?P<qno>\d{3})[/\\](?P<slot>question|option[1-4])-(?P<idx>\d+)"
)


def normalize_nfc(value: str) -> str:
    return unicodedata.normalize("NFC", value)


def select_by_rounds[T](items: list[T], rounds: set[int]) -> list[tuple[int, T]]:
    selected: list[tuple[int, T]] = []
    for idx, item in enumerate(items, start=1):
        if idx in rounds:
            selected.append((idx, item))
    return selected


def detect_input_root(project_root: Path) -> Path:
    candidates = [
        directory
        for directory in project_root.iterdir()
        if directory.is_dir() and normalize_nfc(directory.name) == "실기 시험"
    ]
    if candidates:
        return candidates[0]
    return project_root / "실기 시험"


def discover_target_pdfs(
    project_root: Path,
    year: int,
    rounds: set[int],
    exam_prefix: str,
) -> list[tuple[int, str, Path]]:
    input_root = detect_input_root(project_root)
    if not input_root.exists():
        raise FileNotFoundError(f"입력 폴더를 찾지 못했습니다: {input_root}")

    pattern = re.compile(rf"^{re.escape(exam_prefix)}(?P<date>\d{{8}})\(교사용\)$")
    candidates: list[tuple[str, Path]] = []
    for pdf_path in sorted(input_root.rglob("*.pdf")):
        normalized_stem = normalize_nfc(pdf_path.stem)
        match = pattern.match(normalized_stem)
        if not match:
            continue
        date = match.group("date")
        if not date.startswith(str(year)):
            continue
        candidates.append((date, pdf_path))

    if not candidates:
        raise ValueError(f"{year}년 {exam_prefix} PDF를 찾지 못했습니다. 경로: {input_root}")

    candidates.sort(key=lambda item: item[0])
    selected = select_by_rounds(candidates, rounds)
    if not selected:
        raise ValueError(
            f"{year}년 {exam_prefix} PDF는 찾았지만 요청 회차 {sorted(rounds)}가 없습니다."
        )

    return [(round_no, date, path) for round_no, (date, path) in selected]


def discover_exam_files(
    project_root: Path,
    subject_key: str,
    year: int,
    rounds: set[int],
) -> list[ExamFileSet]:
    excel_dir = project_root / "output" / "excel" / subject_key
    txt_dir = project_root / "output" / "text" / subject_key
    image_root = project_root / "output" / "images" / subject_key

    if not excel_dir.exists():
        raise FileNotFoundError(f"Excel directory not found: {excel_dir}")

    excel_candidates: list[tuple[str, str, Path]] = []
    for excel_path in sorted(excel_dir.glob("*.xlsx")):
        # 파일명에서 날짜 추출
        match = re.search(r"(\d{8})", excel_path.stem)
        if not match:
            continue
        date = match.group(1)
        if not date.startswith(str(year)):
            continue
        excel_candidates.append((date, excel_path.stem, excel_path))

    if not excel_candidates:
        raise ValueError(
            f"{year}년 {subject_key} Excel을 찾지 못했습니다. 현재 경로: {excel_dir}"
        )

    excel_candidates.sort(key=lambda item: item[0])
    discovered: list[ExamFileSet] = []
    for idx, (date, stem, excel_path) in select_by_rounds(excel_candidates, rounds):
        txt_path = txt_dir / f"{stem}.txt"
        image_dir = image_root / stem
        if not txt_path.exists():
            raise FileNotFoundError(f"TXT not found for {stem}: {txt_path}")
        if not image_dir.exists():
            image_dir.mkdir(parents=True, exist_ok=True)
        round_id = f"{year}-{idx}"
        discovered.append(
            ExamFileSet(
                stem=stem,
                date=date,
                csv_path=None,  # CSV는 더 이상 사용하지 않음
                excel_path=excel_path,
                txt_path=txt_path,
                image_dir=image_dir,
                round_no=idx,
                round_id=round_id,
            )
        )

    if not discovered:
        raise ValueError(
            f"{year}년 데이터는 찾았지만 요청한 회차 {sorted(rounds)}에 해당하는 파일이 없습니다."
        )
    return discovered


def load_questions(excel_path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "openpyxl is not installed. Run `pip install -r requirements.txt` first."
        ) from exc

    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    loaded: list[dict[str, str]] = []
    for values in rows[1:]:
        row = {}
        for idx, value in enumerate(values):
            if idx >= len(headers) or not headers[idx]:
                continue
            
            if value is None:
                row[headers[idx]] = ""
            else:
                value_str = str(value).strip()
                # Excel에서 수식 방지를 위해 추가된 작은따옴표 제거
                if value_str.startswith("'") and len(value_str) > 1:
                    value_str = value_str[1:]
                row[headers[idx]] = value_str
        
        number_raw = row.get("번호", "")
        if not number_raw.isdigit():
            continue
        # 번호만 있고 나머지가 빈칸인 템플릿 행은 업로드에서 제외한다.
        payload_fields = ["질문", "문항1", "문항2", "문항3", "문항4", "정답"]
        if not any(row.get(field, "") for field in payload_fields):
            continue
        loaded.append(row)
    return loaded


def build_page_questions_map(txt_path: Path) -> dict[int, list[int]]:
    raw_text = txt_path.read_text(encoding="utf-8")
    pieces = PAGE_MARKER_PATTERN.split(raw_text)
    page_map: dict[int, list[int]] = {}

    # split 결과: [prefix, page_no_1, body_1, page_no_2, body_2, ...]
    for idx in range(1, len(pieces), 2):
        page_no = int(pieces[idx])
        page_body = pieces[idx + 1] if idx + 1 < len(pieces) else ""
        seen: set[int] = set()
        ordered_numbers: list[int] = []
        for match in QUESTION_NUMBER_PATTERN.finditer(page_body):
            number = int(match.group(1))
            if number < 1 or number > 100:
                continue
            if number in seen:
                continue
            seen.add(number)
            ordered_numbers.append(number)
        page_map[page_no] = ordered_numbers
    return page_map


def assign_images_to_questions(
    image_dir: Path,
    page_questions_map: dict[int, list[int]],
) -> tuple[dict[int, list[Path]], list[Path]]:
    """이미지 파일을 문제 번호별로 그룹화합니다.
    
    구버전 형식(page_001_img_01.jpg)과 신규 형식(images/.../001/question-1.jpg) 모두 지원합니다.
    """
    # 신규 형식: 디렉토리 구조로 탐색
    image_files = []
    for path in image_dir.rglob("*"):
        if path.is_file():
            image_files.append(path)
    
    image_files = sorted(image_files)
    question_images: dict[int, list[Path]] = defaultdict(list)
    unassigned: list[Path] = []

    # 신규 형식 먼저 확인 (경로 기반)
    has_new_format = False
    for path in image_files:
        path_str = str(path).replace("\\", "/")
        match = FIREBASE_IMAGE_PATH_PATTERN.search(path_str)
        if match:
            has_new_format = True
            question_no = int(match.group("qno"))
            question_images[question_no].append(path)
    
    # 신규 형식이 있으면 그것을 사용하고, 나머지는 unassigned로 처리
    if has_new_format:
        for path in image_files:
            path_str = str(path).replace("\\", "/")
            match = FIREBASE_IMAGE_PATH_PATTERN.search(path_str)
            if not match:
                unassigned.append(path)
        return dict(question_images), unassigned
    
    # 신규 형식이 없으면 구버전 방식 사용
    grouped_by_page: dict[int, list[tuple[int, Path]]] = defaultdict(list)
    for path in image_files:
        match = IMAGE_NAME_PATTERN.match(path.name)
        if not match:
            unassigned.append(path)
            continue
        page = int(match.group("page"))
        idx = int(match.group("idx"))
        grouped_by_page[page].append((idx, path))

    for page, items in grouped_by_page.items():
        items.sort(key=lambda item: item[0])
        question_numbers = page_questions_map.get(page, [])
        if not question_numbers:
            for _, path in items:
                unassigned.append(path)
            continue
        for zero_based, (_, path) in enumerate(items):
            question_idx = min(zero_based, len(question_numbers) - 1)
            question_no = question_numbers[question_idx]
            question_images[question_no].append(path)

    return dict(question_images), unassigned

