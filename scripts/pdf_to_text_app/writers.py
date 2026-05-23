from __future__ import annotations

import json
from dataclasses import asdict
from pathlib import Path

from .models import EXCEL_IMAGE_COUNT_FIELDS, ExtractionResult, QUESTION_FIELDS


def write_questions_excel(excel_path: Path, questions: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook, load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "openpyxl is not installed. Run `pip install -r requirements.txt` first."
        ) from exc

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    headers = [*QUESTION_FIELDS, *EXCEL_IMAGE_COUNT_FIELDS]
    
    # 기존 Excel 파일이 있으면 수동 편집 내용을 병합
    if excel_path.exists():
        try:
            existing_data = _read_existing_excel(excel_path)
            questions = _merge_excel_data(existing_data, questions, headers)
        except Exception as e:
            print(f"[WARN] 기존 Excel 읽기 실패, 새로 생성합니다: {e}")
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "문제목록"
    sheet.append(headers)
    
    for row in questions:
        row_values = []
        for key in headers:
            value = row.get(key, "")
            # 수식 방지: =, +, -, @로 시작하는 텍스트는 앞에 '를 붙여서 텍스트로 저장
            if value and isinstance(value, str) and value[0] in ('=', '+', '-', '@'):
                value = "'" + value
            row_values.append(value)
        sheet.append(row_values)
    
    workbook.save(excel_path)


def _read_existing_excel(excel_path: Path) -> dict[str, dict[str, str]]:
    """기존 Excel 파일을 읽어서 번호별 데이터를 딕셔너리로 반환합니다."""
    from openpyxl import load_workbook
    
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    
    # 헤더 읽기
    headers = []
    for cell in sheet[1]:
        if cell.value:
            headers.append(str(cell.value))
    
    if not headers:
        return {}
    
    # 데이터 읽기
    existing_data: dict[str, dict[str, str]] = {}
    for row_idx in range(2, sheet.max_row + 1):
        row_data: dict[str, str] = {}
        question_no = None
        
        for col_idx, header in enumerate(headers, start=1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            value = str(cell_value) if cell_value is not None else ""
            
            # Excel에서 수식 방지를 위해 추가된 작은따옴표 제거
            if value.startswith("'") and len(value) > 1:
                value = value[1:]
            
            row_data[header] = value
            
            if header == "번호" and value:
                question_no = value
        
        if question_no:
            existing_data[question_no] = row_data
    
    return existing_data


def _merge_excel_data(
    existing_data: dict[str, dict[str, str]],
    new_questions: list[dict[str, str]],
    headers: list[str],
) -> list[dict[str, str]]:
    """기존 Excel 데이터와 새로 파싱된 데이터를 병합합니다.
    
    병합 규칙:
    1. 기존에 수동으로 편집된 내용(질문, 문항, 정답, 과목)은 보존
    2. 이미지 카운트 필드는 새로 계산된 값으로 업데이트
    3. 기존에 없던 문제는 새로 추가
    """
    merged: list[dict[str, str]] = []
    
    # 수동 편집 필드 (보존할 필드)
    manual_fields = {"번호", "질문", "문항1", "문항2", "문항3", "문항4", "정답", "과목"}
    # 자동 생성 필드 (항상 업데이트할 필드)
    auto_fields = set(EXCEL_IMAGE_COUNT_FIELDS)
    
    for new_row in new_questions:
        question_no = new_row.get("번호", "")
        
        if question_no in existing_data:
            existing_row = existing_data[question_no]
            merged_row: dict[str, str] = {}
            
            for header in headers:
                if header in manual_fields:
                    # 수동 편집 필드: 기존 값이 비어있지 않으면 기존 값 사용
                    existing_value = existing_row.get(header, "").strip()
                    if existing_value:
                        merged_row[header] = existing_value
                    else:
                        merged_row[header] = new_row.get(header, "")
                elif header in auto_fields:
                    # 자동 생성 필드: 항상 새 값 사용
                    merged_row[header] = new_row.get(header, "")
                else:
                    # 기타 필드: 새 값 우선
                    merged_row[header] = new_row.get(header, "")
            
            merged.append(merged_row)
        else:
            # 기존에 없던 문제는 그대로 추가
            merged.append(new_row)
    
    return merged


def prepare_excel_rows(
    questions: list[dict[str, str]],
    expected_count: int | None,
    image_count_map: dict[int, dict[str, int]],
) -> list[dict[str, str]]:
    by_number: dict[int, dict[str, str]] = {}
    for row in questions:
        number_raw = row.get("번호", "")
        if not number_raw.isdigit():
            continue
        by_number[int(number_raw)] = dict(row)

    if expected_count is not None and expected_count > 0:
        target_max = expected_count
    else:
        target_max = max(by_number.keys(), default=0)

    rows: list[dict[str, str]] = []
    for number in range(1, target_max + 1):
        base = by_number.get(
            number,
            {
                "번호": str(number),
                "질문": "",
                "문항1": "",
                "문항2": "",
                "문항3": "",
                "문항4": "",
                "정답": "",
                "과목": "",
            },
        )
        counts = image_count_map.get(number, {})
        base["문제이미지"] = str(counts.get("question", 0))
        base["문항1이미지"] = str(counts.get("option1", 0))
        base["문항2이미지"] = str(counts.get("option2", 0))
        base["문항3이미지"] = str(counts.get("option3", 0))
        base["문항4이미지"] = str(counts.get("option4", 0))
        rows.append(base)
    return rows


def write_manifest(
    manifest_path: Path,
    input_root: Path,
    output_root: Path,
    pdf_files: list[Path],
    results: list[ExtractionResult],
) -> None:
    manifest = {
        "input_dir": str(input_root),
        "output_dir": str(output_root),
        "total_files": len(pdf_files),
        "results": [asdict(item) for item in results],
    }
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

